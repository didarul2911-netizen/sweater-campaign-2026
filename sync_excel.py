"""
Sync & Update Master Excel File (Sweater_Campaign_2026_Master_Format.xlsx)
Run this script anytime to update the Master Excel file on your computer with the latest campaign inputs.
"""
import openpyxl
import json
import os

EXCEL_FILE = "Sweater_Campaign_2026_Master_Format.xlsx"

def update_excel_from_store(store_data):
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # Sheet 1: Gyne Core Doctor (Family)
    ws1 = wb["Gyne Core Doctor (Family)"]
    for row in range(3, ws1.max_row + 1):
        terr_code = str(ws1.cell(row=row, column=5).value) # Col E is SAP Territory Code
        if terr_code in store_data:
            d = store_data[terr_code]
            ws1.cell(row=row, column=7, value=d.get("c1_doc_name", ""))
            ws1.cell(row=row, column=8, value=d.get("c1_m1_sweater", ""))
            ws1.cell(row=row, column=9, value=d.get("c1_m1_size", ""))
            ws1.cell(row=row, column=10, value=d.get("c1_m2_sweater", ""))
            ws1.cell(row=row, column=11, value=d.get("c1_m2_size", ""))
            ws1.cell(row=row, column=12, value=d.get("c1_m3_sweater", ""))
            ws1.cell(row=row, column=13, value=d.get("c1_m3_size", ""))
            ws1.cell(row=row, column=14, value=d.get("c1_m4_sweater", ""))
            ws1.cell(row=row, column=15, value=d.get("c1_m4_size", ""))

    # Sheet 2: Core Doctor Maximization
    ws2 = wb["Core Doctor Maximization"]
    for row in range(3, ws2.max_row + 1):
        terr_code = str(ws2.cell(row=row, column=5).value) # Col E
        if terr_code in store_data:
            d = store_data[terr_code]
            ws2.cell(row=row, column=7, value=d.get("c2_d1_name", ""))
            ws2.cell(row=row, column=8, value=d.get("c2_d1_sweater", ""))
            ws2.cell(row=row, column=9, value=d.get("c2_d1_size", ""))
            ws2.cell(row=row, column=10, value=d.get("c2_d2_name", ""))
            ws2.cell(row=row, column=11, value=d.get("c2_d2_sweater", ""))
            ws2.cell(row=row, column=12, value=d.get("c2_d2_size", ""))
            ws2.cell(row=row, column=13, value=d.get("c2_d3_name", ""))
            ws2.cell(row=row, column=14, value=d.get("c2_d3_sweater", ""))
            ws2.cell(row=row, column=15, value=d.get("c2_d3_size", ""))
            ws2.cell(row=row, column=16, value=d.get("c2_d4_name", ""))
            ws2.cell(row=row, column=17, value=d.get("c2_d4_sweater", ""))
            ws2.cell(row=row, column=18, value=d.get("c2_d4_size", ""))

    wb.save(EXCEL_FILE)
    print(f"Successfully synced and updated {EXCEL_FILE} with all inputs!")

if __name__ == "__main__":
    if os.path.exists("saved_inputs.json"):
        with open("saved_inputs.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        update_excel_from_store(data)
    else:
        print("Ready to sync. When data is exported or saved, run this script to update Master Excel.")
