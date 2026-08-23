import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_updated_master_workbook():
    base_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd()
    ff_path = os.path.join(base_dir, 'FF list.xlsx')
    output_path = os.path.join(base_dir, 'Sweater_Campaign_2026_Master_Format.xlsx')

    print(f"Reading FF List from: {ff_path}")
    df_ff = pd.read_excel(ff_path)
    total_territories = len(df_ff)
    max_row = total_territories + 2
    print(f"Loaded {total_territories} territories.")

    wb = openpyxl.Workbook()
    # Remove default sheet later
    
    # Common Styles
    font_main_title = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_sub_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_bold_data = Font(name="Segoe UI", size=10, bold=True)
    
    fill_territory_group = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Navy
    fill_territory_sub = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")   # Steel Blue
    
    fill_c1_group = PatternFill(start_color="006D77", end_color="006D77", fill_type="solid")        # Teal
    fill_c1_sub = PatternFill(start_color="009688", end_color="009688", fill_type="solid")          # Emerald Teal
    fill_c1_sub_alt = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
    
    fill_c2_group = PatternFill(start_color="4A154B", end_color="4A154B", fill_type="solid")        # Aubergine / Purple
    fill_c2_sub = PatternFill(start_color="6B2D5C", end_color="6B2D5C", fill_type="solid")
    fill_c2_sub_alt = PatternFill(start_color="8338EC", end_color="8338EC", fill_type="solid")
    
    fill_status_sub = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )
    header_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF")
    )

    # -------------------------------------------------------------
    # 1. SHEET: Dropdown_Lists
    # -------------------------------------------------------------
    ws_dropdown = wb.create_sheet(title="Dropdown_Lists")
    ws_dropdown.sheet_state = 'hidden'
    
    dropdown_data = {
        "Sweater_Codes": [
            "01 - Men's V-Neck (Grey)",
            "02 - Men's V-Neck (Navy Blue)",
            "03 - Men's V-Neck (Cream Check)",
            "04 - Women's Short Cardigan (Check)",
            "05 - Women's Semi Long Cardigan (Black)"
        ],
        "Sizes": ["XS", "S", "M", "L", "XL", "XXL"]
    }
    
    ws_dropdown.cell(row=1, column=1, value="Sweater_Codes").font = Font(bold=True)
    for idx, val in enumerate(dropdown_data["Sweater_Codes"], start=2):
        ws_dropdown.cell(row=idx, column=1, value=val)
        
    ws_dropdown.cell(row=1, column=2, value="Sizes").font = Font(bold=True)
    for idx, val in enumerate(dropdown_data["Sizes"], start=2):
        ws_dropdown.cell(row=idx, column=2, value=val)

    # Setup openpyxl Data Validations
    dv_sweater = DataValidation(type="list", formula1="=Dropdown_Lists!$A$2:$A$6", allow_blank=True)
    dv_sweater.error = 'Please select a valid Sweater from the list.'
    dv_size = DataValidation(type="list", formula1="=Dropdown_Lists!$B$2:$B$7", allow_blank=True)
    dv_size.error = 'Please select a valid Size.'

    # -------------------------------------------------------------
    # 2. SHEET 1: Gyne Core Doctor (Family) - Campaign 1
    # -------------------------------------------------------------
    ws_c1 = wb.create_sheet(title="Gyne Core Doctor (Family)", index=0)
    ws_c1.views.sheetView[0].showGridLines = True
    
    # Main Header Row 1
    ws_c1.merge_cells("A1:F1")
    ws_c1["A1"] = "TERRITORY INFORMATION (EXIUM FIELD FORCE LIST)"
    for c in range(1, 7):
        cell = ws_c1.cell(row=1, column=c)
        cell.fill = fill_territory_group
        cell.font = font_main_title
        cell.alignment = align_center
        cell.border = header_border
        
    ws_c1.merge_cells("G1:O1")
    ws_c1["G1"] = "CAMPAIGN 1: GYNE CORE DOCTOR DEVELOPMENT (FAMILY PACKAGE - 4 SWEATERS / TERRITORY)"
    for c in range(7, 16):
        cell = ws_c1.cell(row=1, column=c)
        cell.fill = fill_c1_group
        cell.font = font_main_title
        cell.alignment = align_center
        cell.border = header_border

    ws_c1.merge_cells("P1:P1")
    ws_c1["P1"] = "STATUS"
    ws_c1["P1"].fill = fill_status_sub
    ws_c1["P1"].font = font_main_title
    ws_c1["P1"].alignment = align_center
    ws_c1["P1"].border = header_border
    
    # Sub-Headers Row 2
    c1_headers = [
        ("Zone", fill_territory_sub),
        ("SAP Region Code", fill_territory_sub),
        ("Region", fill_territory_sub),
        ("Regional Head", fill_territory_sub),
        ("SAP Territory Code", fill_territory_sub),
        ("Territory", fill_territory_sub),
        ("Doctor Name", fill_c1_sub),
        ("Sweater 1", fill_c1_sub_alt),
        ("Size 1", fill_c1_sub_alt),
        ("Sweater 2", fill_c1_sub),
        ("Size 2", fill_c1_sub),
        ("Sweater 3", fill_c1_sub_alt),
        ("Size 3", fill_c1_sub_alt),
        ("Sweater 4", fill_c1_sub),
        ("Size 4", fill_c1_sub),
        ("Territory Status", fill_status_sub)
    ]
    
    ws_c1.row_dimensions[1].height = 28
    ws_c1.row_dimensions[2].height = 30
    
    for c_idx, (h_text, fill_c) in enumerate(c1_headers, start=1):
        cell = ws_c1.cell(row=2, column=c_idx, value=h_text)
        cell.fill = fill_c
        cell.font = font_sub_header
        cell.alignment = align_center
        cell.border = header_border

    # Insert Territory Rows for C1
    for row_idx, r in df_ff.iterrows():
        e_row = row_idx + 3
        ws_c1.cell(row=e_row, column=1, value=r['Zone'])
        ws_c1.cell(row=e_row, column=2, value=r['SAP Region Code'])
        ws_c1.cell(row=e_row, column=3, value=r['Region'])
        ws_c1.cell(row=e_row, column=4, value=r['Regional Head'])
        ws_c1.cell(row=e_row, column=5, value=r['SAP Territory Code'])
        ws_c1.cell(row=e_row, column=6, value=r['Territory'])
        
        for c_idx in range(1, 7):
            c = ws_c1.cell(row=e_row, column=c_idx)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_center if c_idx in [2, 5] else align_left

        for c_idx in range(7, 16):
            c = ws_c1.cell(row=e_row, column=c_idx)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_center if c_idx in [8, 9, 10, 11, 12, 13, 14, 15] else align_left

        # Status Formula
        status_f = f'=IF(AND(G{e_row}<>"",H{e_row}<>"",I{e_row}<>"",J{e_row}<>"",K{e_row}<>"",L{e_row}<>"",M{e_row}<>"",N{e_row}<>"",O{e_row}<>""), "Complete", IF(OR(G{e_row}<>"",H{e_row}<>"",J{e_row}<>"",L{e_row}<>"",N{e_row}<>""), "In Progress", "Not Started"))'
        sc = ws_c1.cell(row=e_row, column=16, value=status_f)
        sc.font = font_bold_data
        sc.alignment = align_center
        sc.border = thin_border

    # Validations for C1
    ws_c1.add_data_validation(dv_sweater)
    ws_c1.add_data_validation(dv_size)
    for col_l in ['H', 'J', 'L', 'N']:
        dv_sweater.add(f"{col_l}3:{col_l}{max_row}")
    for col_l in ['I', 'K', 'M', 'O']:
        dv_size.add(f"{col_l}3:{col_l}{max_row}")

    ws_c1.freeze_panes = "G3"
    c1_widths = {'A': 18, 'B': 16, 'C': 20, 'D': 22, 'E': 18, 'F': 20, 'G': 26, 'H': 32, 'I': 12, 'J': 32, 'K': 12, 'L': 32, 'M': 12, 'N': 32, 'O': 12, 'P': 16}
    for col_l, w in c1_widths.items():
        ws_c1.column_dimensions[col_l].width = w

    # -------------------------------------------------------------
    # 3. SHEET 2: Core Doctor Maximization - Campaign 2
    # -------------------------------------------------------------
    ws_c2 = wb.create_sheet(title="Core Doctor Maximization", index=1)
    ws_c2.views.sheetView[0].showGridLines = True
    
    # Main Header Row 1
    ws_c2.merge_cells("A1:F1")
    ws_c2["A1"] = "TERRITORY INFORMATION (EXIUM FIELD FORCE LIST)"
    for c in range(1, 7):
        cell = ws_c2.cell(row=1, column=c)
        cell.fill = fill_territory_group
        cell.font = font_main_title
        cell.alignment = align_center
        cell.border = header_border
        
    ws_c2.merge_cells("G1:R1")
    ws_c2["G1"] = "CAMPAIGN 2: CORE DOCTOR MAXIMIZATION (1 SWEATER / DOCTOR - 4 DOCTORS / TERRITORY)"
    for c in range(7, 19):
        cell = ws_c2.cell(row=1, column=c)
        cell.fill = fill_c2_group
        cell.font = font_main_title
        cell.alignment = align_center
        cell.border = header_border

    ws_c2.merge_cells("S1:S1")
    ws_c2["S1"] = "STATUS"
    ws_c2["S1"].fill = fill_status_sub
    ws_c2["S1"].font = font_main_title
    ws_c2["S1"].alignment = align_center
    ws_c2["S1"].border = header_border
    
    # Sub-Headers Row 2
    c2_headers = [
        ("Zone", fill_territory_sub),
        ("SAP Region Code", fill_territory_sub),
        ("Region", fill_territory_sub),
        ("Regional Head", fill_territory_sub),
        ("SAP Territory Code", fill_territory_sub),
        ("Territory", fill_territory_sub),
        ("Doctor 1 Name", fill_c2_sub),
        ("Sweater 1", fill_c2_sub),
        ("Size 1", fill_c2_sub),
        ("Doctor 2 Name", fill_c2_sub_alt),
        ("Sweater 2", fill_c2_sub_alt),
        ("Size 2", fill_c2_sub_alt),
        ("Doctor 3 Name", fill_c2_sub),
        ("Sweater 3", fill_c2_sub),
        ("Size 3", fill_c2_sub),
        ("Doctor 4 Name", fill_c2_sub_alt),
        ("Sweater 4", fill_c2_sub_alt),
        ("Size 4", fill_c2_sub_alt),
        ("Territory Status", fill_status_sub)
    ]
    
    ws_c2.row_dimensions[1].height = 28
    ws_c2.row_dimensions[2].height = 30
    
    for c_idx, (h_text, fill_c) in enumerate(c2_headers, start=1):
        cell = ws_c2.cell(row=2, column=c_idx, value=h_text)
        cell.fill = fill_c
        cell.font = font_sub_header
        cell.alignment = align_center
        cell.border = header_border

    # Insert Territory Rows for C2
    for row_idx, r in df_ff.iterrows():
        e_row = row_idx + 3
        ws_c2.cell(row=e_row, column=1, value=r['Zone'])
        ws_c2.cell(row=e_row, column=2, value=r['SAP Region Code'])
        ws_c2.cell(row=e_row, column=3, value=r['Region'])
        ws_c2.cell(row=e_row, column=4, value=r['Regional Head'])
        ws_c2.cell(row=e_row, column=5, value=r['SAP Territory Code'])
        ws_c2.cell(row=e_row, column=6, value=r['Territory'])
        
        for c_idx in range(1, 7):
            c = ws_c2.cell(row=e_row, column=c_idx)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_center if c_idx in [2, 5] else align_left

        for c_idx in range(7, 19):
            c = ws_c2.cell(row=e_row, column=c_idx)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_center if c_idx in [8, 9, 11, 12, 14, 15, 17, 18] else align_left

        # Status Formula
        status_f2 = f'=IF(AND(G{e_row}<>"",H{e_row}<>"",I{e_row}<>"",J{e_row}<>"",K{e_row}<>"",L{e_row}<>"",M{e_row}<>"",N{e_row}<>"",O{e_row}<>"",P{e_row}<>"",Q{e_row}<>"",R{e_row}<>""), "Complete", IF(OR(G{e_row}<>"",J{e_row}<>"",M{e_row}<>"",P{e_row}<>""), "In Progress", "Not Started"))'
        sc2 = ws_c2.cell(row=e_row, column=19, value=status_f2)
        sc2.font = font_bold_data
        sc2.alignment = align_center
        sc2.border = thin_border

    # Validations for C2
    dv_sweater_c2 = DataValidation(type="list", formula1="=Dropdown_Lists!$A$2:$A$6", allow_blank=True)
    dv_size_c2 = DataValidation(type="list", formula1="=Dropdown_Lists!$B$2:$B$7", allow_blank=True)
    ws_c2.add_data_validation(dv_sweater_c2)
    ws_c2.add_data_validation(dv_size_c2)
    for col_l in ['H', 'K', 'N', 'Q']:
        dv_sweater_c2.add(f"{col_l}3:{col_l}{max_row}")
    for col_l in ['I', 'L', 'O', 'R']:
        dv_size_c2.add(f"{col_l}3:{col_l}{max_row}")

    ws_c2.freeze_panes = "G3"
    c2_widths = {'A': 18, 'B': 16, 'C': 20, 'D': 22, 'E': 18, 'F': 20, 'G': 24, 'H': 32, 'I': 12, 'J': 24, 'K': 32, 'L': 12, 'M': 24, 'N': 32, 'O': 12, 'P': 24, 'Q': 32, 'R': 12, 'S': 16}
    for col_l, w in c2_widths.items():
        ws_c2.column_dimensions[col_l].width = w

    # -------------------------------------------------------------
    # 4. SHEET 3: Procurement_Summary
    # -------------------------------------------------------------
    ws_summary = wb.create_sheet(title="Procurement_Summary", index=2)
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title
    ws_summary.merge_cells("A1:H1")
    ws_summary["A1"] = "EXIUM 4Q'26 SWEATER CAMPAIGN - TOTAL PROCUREMENT REQUIREMENT"
    ws_summary["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws_summary["A1"].alignment = align_center
    ws_summary.row_dimensions[1].height = 35

    ws_summary.merge_cells("A2:H2")
    ws_summary["A2"] = "Live automated aggregation across Campaign 1 (Family Package) and Campaign 2 (Core Doctor Maximization)"
    ws_summary["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="555555")
    ws_summary["A2"].alignment = align_left
    ws_summary.row_dimensions[2].height = 20

    # Table Header (Row 4)
    matrix_headers = ["Sweater Design & Description", "XS", "S", "M", "L", "XL", "XXL", "TOTAL ORDER"]
    ws_summary.row_dimensions[4].height = 26
    for col_idx, h in enumerate(matrix_headers, start=1):
        cell = ws_summary.cell(row=4, column=col_idx, value=h)
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.alignment = align_center
        cell.border = header_border

    sweaters = [
        ("01 - Men's V-Neck (Grey)", "Men's Sleeveless V-Neck (Ash / Grey Textured)"),
        ("02 - Men's V-Neck (Navy Blue)", "Men's Sleeveless V-Neck (Navy Blue Textured)"),
        ("03 - Men's V-Neck (Cream Check)", "Men's Sleeveless V-Neck (Off-White / Cream Checkered)"),
        ("04 - Women's Short Cardigan (Check)", "Women's Short Cardigan (White & Navy Checkered Button-Up)"),
        ("05 - Women's Semi Long Cardigan (Black)", "Women's Semi Long Cardigan (Black with Border Trim)")
    ]

    c1_sw_cols = ['H', 'J', 'L', 'N']
    c1_sz_cols = ['I', 'K', 'M', 'O']

    c2_sw_cols = ['H', 'K', 'N', 'Q']
    c2_sz_cols = ['I', 'L', 'O', 'R']

    size_headers = ["XS", "S", "M", "L", "XL", "XXL"]

    for row_offset, (sw_code, sw_desc) in enumerate(sweaters, start=5):
        ws_summary.row_dimensions[row_offset].height = 24
        cell_name = ws_summary.cell(row=row_offset, column=1, value=sw_code)
        cell_name.font = font_bold_data
        cell_name.border = thin_border
        cell_name.alignment = align_left
        
        for col_offset, sz in enumerate(size_headers, start=2):
            col_letter = get_column_letter(col_offset)
            
            # C1 countifs
            c1_terms = [f"'Gyne Core Doctor (Family)'!${sc}$3:${sc}${max_row}, $A{row_offset}, 'Gyne Core Doctor (Family)'!${szc}$3:${szc}${max_row}, {col_letter}$4"
                        for sc, szc in zip(c1_sw_cols, c1_sz_cols)]
            
            # C2 countifs
            c2_terms = [f"'Core Doctor Maximization'!${sc}$3:${sc}${max_row}, $A{row_offset}, 'Core Doctor Maximization'!${szc}$3:${szc}${max_row}, {col_letter}$4"
                        for sc, szc in zip(c2_sw_cols, c2_sz_cols)]
            
            all_terms = [f"COUNTIFS({t})" for t in (c1_terms + c2_terms)]
            formula = "=" + " + ".join(all_terms)
            
            cell_val = ws_summary.cell(row=row_offset, column=col_offset, value=formula)
            cell_val.font = font_data
            cell_val.alignment = align_center
            cell_val.border = thin_border
            
        tot_cell = ws_summary.cell(row=row_offset, column=8, value=f"=SUM(B{row_offset}:G{row_offset})")
        tot_cell.font = font_bold_data
        tot_cell.alignment = align_center
        tot_cell.border = thin_border
        tot_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Grand Total Row
    ws_summary.row_dimensions[10].height = 26
    tot_label = ws_summary.cell(row=10, column=1, value="GRAND TOTAL PIECES")
    tot_label.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    tot_label.fill = PatternFill(start_color="137547", end_color="137547", fill_type="solid")
    tot_label.alignment = align_left
    tot_label.border = header_border

    for col_offset in range(2, 8):
        col_letter = get_column_letter(col_offset)
        c = ws_summary.cell(row=10, column=col_offset, value=f"=SUM({col_letter}5:{col_letter}9)")
        c.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="137547", end_color="137547", fill_type="solid")
        c.alignment = align_center
        c.border = header_border

    gt_cell = ws_summary.cell(row=10, column=8, value=f"=SUM(H5:H9)")
    gt_cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    gt_cell.fill = PatternFill(start_color="137547", end_color="137547", fill_type="solid")
    gt_cell.alignment = align_center
    gt_cell.border = header_border

    # KPI Targets Card
    ws_summary.cell(row=12, column=1, value="CAMPAIGN TARGET METRICS").font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    kpis = [
        ("Total Active Territories", total_territories),
        ("Sweaters / Territory (Campaign 1: Gyne Core Family)", 4),
        ("Sweaters / Territory (Campaign 2: Core Doctor Max.)", 4),
        ("Total Sweaters / Territory", 8),
        ("Total Expected Procurement Quantity", total_territories * 8),
        ("Total Expected Participating Doctors", total_territories * 5)
    ]
    for idx, (label, val) in enumerate(kpis, start=13):
        ws_summary.cell(row=idx, column=1, value=label).font = font_data
        ws_summary.cell(row=idx, column=1).border = thin_border
        c_v = ws_summary.cell(row=idx, column=2, value=val)
        c_v.font = font_bold_data
        c_v.alignment = align_center
        c_v.border = thin_border
        c_v.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        ws_summary.merge_cells(f"C{idx}:D{idx}")
        if idx == 17:
            ws_summary.cell(row=idx, column=3, value=f"(= {total_territories} x 8)").font = Font(italic=True, size=9)
        elif idx == 18:
            ws_summary.cell(row=idx, column=3, value=f"(= {total_territories} x 5)").font = Font(italic=True, size=9)

    ws_summary.column_dimensions['A'].width = 42
    ws_summary.column_dimensions['B'].width = 14
    ws_summary.column_dimensions['C'].width = 14
    ws_summary.column_dimensions['D'].width = 14
    ws_summary.column_dimensions['E'].width = 14
    ws_summary.column_dimensions['F'].width = 14
    ws_summary.column_dimensions['G'].width = 14
    ws_summary.column_dimensions['H'].width = 18

    # -------------------------------------------------------------
    # 5. SHEET 4: Product_Catalog_&_Size_Chart
    # -------------------------------------------------------------
    ws_catalog = wb.create_sheet(title="Product_Catalog_&_Size_Chart", index=3)
    ws_catalog.views.sheetView[0].showGridLines = True
    
    ws_catalog.merge_cells("A1:F1")
    ws_catalog["A1"] = "PRODUCT CATALOG & SPECIFICATION GUIDE - 4Q'26 SWEATER CAMPAIGN"
    ws_catalog["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    ws_catalog["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws_catalog["A1"].alignment = align_center
    ws_catalog.row_dimensions[1].height = 35

    cat_headers = ["Code", "Category / Gender", "Style & Model Name", "Color / Pattern", "Available Sizes", "Supplier / Brand"]
    ws_catalog.row_dimensions[3].height = 25
    for col_idx, h in enumerate(cat_headers, start=1):
        cell = ws_catalog.cell(row=3, column=col_idx, value=h)
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.alignment = align_center
        cell.border = header_border

    cat_items = [
        ("01", "Men's", "Men's Sleeveless V-Neck Sweater", "Solid Ash / Grey Textured", "S, M, L, XL, XXL", "Richman / Lubnan"),
        ("02", "Men's", "Men's Sleeveless V-Neck Sweater", "Solid Navy Blue Textured", "S, M, L, XL, XXL", "Richman / Lubnan"),
        ("03", "Men's", "Men's Sleeveless V-Neck Sweater", "Cream / Off-White Checkered", "S, M, L, XL, XXL", "Richman / Lubnan"),
        ("04", "Women's", "Short Cardigan Sweater (Button-up)", "White & Navy Grid Check", "XS, S, M, L, XL", "Richman / Lubnan"),
        ("05", "Women's", "Semi Long Cardigan Sweater", "Solid Black with Tribal Border Trim", "S, M, L, XL, XXL", "Richman / Lubnan")
    ]
    for r_idx, item in enumerate(cat_items, start=4):
        ws_catalog.row_dimensions[r_idx].height = 24
        for c_idx, val in enumerate(item, start=1):
            c = ws_catalog.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data
            c.border = thin_border
            if c_idx in [1, 2, 5]:
                c.alignment = align_center
            else:
                c.alignment = align_left

    # Size Charts
    ws_catalog.cell(row=11, column=1, value="1. MEN'S SWEATER (HALF SLEEVE / SLEEVELESS) - MEASUREMENT IN CM").font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    men_size_headers = ["SL", "Measurement Point (CM)", "S", "M", "L", "XL", "XXL"]
    ws_catalog.row_dimensions[12].height = 24
    for c_idx, h in enumerate(men_size_headers, start=1):
        c = ws_catalog.cell(row=12, column=c_idx, value=h)
        c.fill = PatternFill(start_color="137547", end_color="137547", fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = align_center
        c.border = header_border

    men_size_rows = [
        (1, "BODY LENGTH", 65, 67, 69, 71, 73),
        (2, "1/2 CHEST", 48, 50, 52, 54, 56)
    ]
    for r_idx, srow in enumerate(men_size_rows, start=13):
        ws_catalog.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(srow, start=1):
            c = ws_catalog.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_left if c_idx == 2 else align_center

    ws_catalog.cell(row=17, column=1, value="2. WOMEN'S SHORT CARDIGAN SWEATER (STYLE 04) - MEASUREMENT IN CM").font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    w_short_headers = ["SL", "Measurement Point (CM)", "XS", "S", "M", "L", "XL"]
    ws_catalog.row_dimensions[18].height = 24
    for c_idx, h in enumerate(w_short_headers, start=1):
        c = ws_catalog.cell(row=18, column=c_idx, value=h)
        c.fill = PatternFill(start_color="7B2CBF", end_color="7B2CBF", fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = align_center
        c.border = header_border

    w_short_rows = [
        (1, "BODY LENGTH FROM HPS", 60, 62, 64, 66, 68),
        (2, "1/2 CHEST", 44, 46, 48, 50, 52),
        (3, "LONG SLEEVE LENGTH", 57, 58, 59, 60, 61)
    ]
    for r_idx, srow in enumerate(w_short_rows, start=19):
        ws_catalog.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(srow, start=1):
            c = ws_catalog.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_left if c_idx == 2 else align_center

    ws_catalog.cell(row=24, column=1, value="3. WOMEN'S SEMI LONG CARDIGAN SWEATER (STYLE 05) - MEASUREMENT IN CM").font = Font(name="Segoe UI", size=11, bold=True, color="1F4E79")
    w_long_headers = ["SL", "Measurement Point (CM)", "S", "M", "L", "XL", "XXL"]
    ws_catalog.row_dimensions[25].height = 24
    for c_idx, h in enumerate(w_long_headers, start=1):
        c = ws_catalog.cell(row=25, column=c_idx, value=h)
        c.fill = PatternFill(start_color="5A189A", end_color="5A189A", fill_type="solid")
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.alignment = align_center
        c.border = header_border

    w_long_rows = [
        (1, "BODY LENGTH FROM HPS", 64, 66, 68, 70, 72),
        (2, "1/2 CHEST", 51, 53, 55, 57, 59),
        (3, "SLEEVE LENGTH", 52, 53, 54, 55, 56)
    ]
    for r_idx, srow in enumerate(w_long_rows, start=26):
        ws_catalog.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(srow, start=1):
            c = ws_catalog.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data
            c.border = thin_border
            c.alignment = align_left if c_idx == 2 else align_center

    ws_catalog.column_dimensions['A'].width = 10
    ws_catalog.column_dimensions['B'].width = 24
    ws_catalog.column_dimensions['C'].width = 38
    ws_catalog.column_dimensions['D'].width = 34
    ws_catalog.column_dimensions['E'].width = 22
    ws_catalog.column_dimensions['F'].width = 24

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    print(f"Saving updated Master Workbook to: {output_path}")
    wb.save(output_path)
    print("Master Workbook successfully generated with 2 distinct campaign sheets!")

if __name__ == '__main__':
    create_updated_master_workbook()
